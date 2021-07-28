import requests
import json
import openpyxl
import datetime
import re
url='https://gql.tokopedia.com/'

# sid 729296 for gudang listrik
def get_base(sid):
    page = 0
    flag = True
    name = []
    product_url = []
    product_id = []
    sold = []
    while flag == True:
        page = page + 1
        paramData = [{"operationName": "ShopProducts",
                      "variables": {"sid": str(sid), "page": page, "perPage": 80, "keyword": "", "etalaseId": "etalase",
                                    "sort": 1},
                      "query": "query ShopProducts($sid: String!, $page: Int, $perPage: Int, $keyword: String, $etalaseId: String, $sort: Int) {\n  GetShopProduct(shopID: $sid, filter: {page: $page, perPage: $perPage, fkeyword: $keyword, fmenu: $etalaseId, sort: $sort}) {\n    status\n    errors\n    links {\n      prev\n      next\n      __typename\n    }\n    data {\n      name\n      product_url\n      product_id\n      price {\n        text_idr\n        __typename\n      }\n      primary_image {\n        original\n        thumbnail\n        resize300\n        __typename\n      }\n      flags {\n        isSold\n        isPreorder\n        isWholesale\n        isWishlist\n        __typename\n      }\n      campaign {\n        discounted_percentage\n        original_price_fmt\n        start_date\n        end_date\n        __typename\n      }\n      label {\n        color_hex\n        content\n        __typename\n      }\n      label_groups {\n        position\n        title\n        type\n        __typename\n      }\n      badge {\n        title\n        image_url\n        __typename\n      }\n      stats {\n        reviewCount\n        rating\n        __typename\n      }\n      category {\n        id\n        __typename\n      }\n      __typename\n    }\n    __typename\n  }\n}\n"}
                     ]
        r = requests.post(
            url=url,
            data=json.dumps(paramData, indent=2)
        )
        k = r.json()
        for i in k[0]['data']['GetShopProduct']['data']:
            name.append(i['name'])
            product_url.append(i['product_url'])
            product_id.append(i['product_id'])
            try:
                sold.append(i['label_groups'][0]['title'])
            except:
                sold.append(None)
        if k[0]['data']['GetShopProduct']['links']['next'] == "":
            flag = False

    shopDomain=re.match('https://www.tokopedia.com/(.*)/(.*)', product_url[0])[1]

    wb = openpyxl.Workbook()

    # grab the active worksheet
    ws = wb.active
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 70

    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws['B2'] = shopDomain
    ws['C3'] = 'times updated'
    ws['D2'] = 'Default = 0'
    ws['D3'] = 0
    ws['A8'] = 'Product ID'
    ws['B8'] = 'Product Name'
    ws['C7'] = 'Date'
    ws['D8'] = datetime.datetime.now()
    ws['C8'].number_format = 'd-mmm-yy'
    ws['D8'].number_format = 'd-mmm-yy'
    for i in range(len(product_url)):
        ws.cell(column=1, row=9 + i, value=int(product_id[i]))
        ws.cell(column=2, row=9 + i, value=re.match('https://www.tokopedia.com/(.*)/(.*)', product_url[i])[2])

    wb.save("sale_report_"+shopDomain+".xlsx")
    wb.save("price_report_"+shopDomain+".xlsx")

get_base(1707404)